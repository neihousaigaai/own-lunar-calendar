from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import get_data


class CalendarPrinter:
	DAYS_OF_WEEK_LABEL = ["T2", "T3", "T4", "T5", "T6", "T7", "CN"]

	MAX_MONTHS_IN_ROW = 3
	MAX_MONTHS_IN_ROW_PORTRAIT = 1

	# style
	month_year_title = Font(name="Comic Sans MS", size=30)

	weekday_ft_color = "000000"
	sat_ft_color = "0000FF"
	sun_ft_color = "FF0000"

	solar_ft_weekday = Font(name="Comic Sans MS", size=24, color=weekday_ft_color)
	solar_ft_sat = Font(name="Comic Sans MS", size=24, color=sat_ft_color)
	solar_ft_sun = Font(name="Comic Sans MS", size=24, color=sun_ft_color)  # font for Sunday/day off cell

	lunar_ft_normal = Font(size=14, color=weekday_ft_color)
	lunar_ft_dayoff = Font(size=14, color=sun_ft_color)  # font for day off cell

	line = Side(border_style="thin", color="000000")
	upper_border = Border(top=line, left=line, right=line)
	lower_border = Border(bottom=line, left=line, right=line)
	left_align = Alignment(horizontal="left", vertical="center")
	right_align = Alignment(horizontal="right", vertical="center")
	center_align = Alignment(horizontal="center", vertical="center")

	def __init__(self, year, init_col, init_row, cnt_each):
		self.days_off = CalendarPrinter.read_days_off()
		self.year = year
		self.init_col = init_col
		self.init_row = init_row
		self.cnt_each = cnt_each

	def get_max_month_in_row(self, orientation):
		if orientation == "landscape":
			return CalendarPrinter.MAX_MONTHS_IN_ROW

		return CalendarPrinter.MAX_MONTHS_IN_ROW_PORTRAIT

	@staticmethod
	def read_days_off():
		fi = open("days_off.txt", "r")
		return [x.rstrip("\n") for x in fi.readlines() if x != ""]

	def is_dayoff(self, date, month, lunar):
		date_format = "{}/{}".format(date, month) + (" (AL)" if lunar else "")
		date_with_year_format = "{}/{}/{}".format(date, month, self.year) + (" (AL)" if lunar else "")
		return date_format in self.days_off or date_with_year_format in self.days_off

	def print_a_year(self, is_online_mode, filename, papersize="A4", orientation="landscape"):
		wb = Workbook()
		sheet = wb.active
		sheet.title = "Sheet1"

		max_month_in_row = self.get_max_month_in_row(orientation)

		for month in range(1, 12 + 1):
			if (month - 1) % self.cnt_each == 0:
				if month > 1:
					sheet = wb.create_sheet("Sheet" + str(month // self.cnt_each + (self.cnt_each > 1)))

				sheet.set_printer_settings(getattr(Worksheet, "PAPERSIZE_" + papersize.upper()), orientation)
				sheet.page_margins.left = 0.393700787  # 1cm
				sheet.page_margins.right = 0.393700787  # 1cm
				sheet.page_margins.top = 0.393700787  # 1cm
				sheet.page_margins.bottom = 0.393700787  # 1cm
				sheet.page_margins.header = 0.125
				sheet.page_margins.footer = 0.125

				sheet.page_setup.fitToPage = True
				sheet.print_options.horizontalCentered = True
				# sheet.print_options.verticalCentered = True

				for col_id in range(1, 8 * min(self.cnt_each, max_month_in_row)):
					# unit: number of characters
					# number of pixels = 7 * number of characters (ref: https://stackoverflow.com/a/63271915/)
					sheet.column_dimensions[get_column_letter(col_id)].width = 15

			month_relative_idx = (month - 1) % self.cnt_each
			if self.cnt_each > max_month_in_row:
				self.print_month(
					sheet,
					self.year, month,
					self.init_col + 8 * (month_relative_idx % max_month_in_row),
					self.init_row + 15 * (month_relative_idx // max_month_in_row),
					is_online_mode
				)
			else:
				self.print_month(
					sheet,
					self.year, month,
					self.init_col + 8 * month_relative_idx,
					self.init_row,
					is_online_mode
				)

		wb.save(filename)

	def print_month(self, sheet, year, month, f_col, f_row, is_online):
		if is_online:
			lst = get_data.crawl_month_amlich(year, month)
		else:
			lst = get_data.get_month_offline(year, month)

		# header
		month_cell = sheet.cell(column=f_col, row=f_row)
		month_cell.value = "{0:0>2}".format(month)
		month_cell.font = CalendarPrinter.month_year_title
		month_cell.alignment = CalendarPrinter.left_align

		year_cell = sheet.cell(column=f_col + 6, row=f_row)
		year_cell.value = str(year)
		year_cell.font = CalendarPrinter.month_year_title
		year_cell.alignment = CalendarPrinter.right_align

		for i in range(7):
			day_cell = sheet.cell(column=f_col + i, row=f_row + 1)
			day_cell.value = CalendarPrinter.DAYS_OF_WEEK_LABEL[i]

			if i == 5:
				font_color = CalendarPrinter.sat_ft_color
			elif i == 6:
				font_color = CalendarPrinter.sun_ft_color
			else:
				font_color = CalendarPrinter.weekday_ft_color

			day_cell.font = Font(size=14, color=font_color)
			day_cell.border = CalendarPrinter.upper_border
			day_cell.alignment = CalendarPrinter.center_align

		l_month = ''
		prev_day_coor = (0, 0)

		for i in range(len(lst)):
			for j in range(7):  # seven days of week
				upper_cell = sheet.cell(column=f_col + j, row=f_row + 2 + 2 * i)
				lower_cell = sheet.cell(column=f_col + j, row=f_row + 2 + 2 * i + 1)

				# set style
				upper_cell.border = CalendarPrinter.upper_border
				lower_cell.border = CalendarPrinter.lower_border

				upper_cell.alignment = CalendarPrinter.left_align
				lower_cell.alignment = CalendarPrinter.right_align

				# assign data
				if lst[i][j] is not None:
					solar_date_str, lunar_date_str = lst[i][j]

					upper_cell.value = solar_date_str
					lower_cell.value = lunar_date_str

					# default state
					if j == 5:
						upper_cell.font = CalendarPrinter.solar_ft_sat
					elif j == 6:
						upper_cell.font = CalendarPrinter.solar_ft_sun
					else:
						upper_cell.font = CalendarPrinter.solar_ft_weekday

					lower_cell.font = CalendarPrinter.lunar_ft_normal

					# new lunar/solar month
					slash_idx = lunar_date_str.find('/')
					if slash_idx != -1:
						l_date = lunar_date_str[:slash_idx]
						if lunar_date_str.find('(') != -1:  # e.g: 1/3 (D/T)
							l_month = lunar_date_str[slash_idx + 1:lunar_date_str.find(' ')]
						else:
							l_month = lunar_date_str[slash_idx + 1:]
					else:
						l_date = lunar_date_str

					# solar date is a day off
					if self.is_dayoff(solar_date_str, month, False):
						upper_cell.font = CalendarPrinter.solar_ft_sun

					# lunar date is a day off
					elif self.is_dayoff(l_date, l_month, True):
						if l_date == '30' and l_month == '12':  # New Year's Eve on 30/12
							prev_lunar_cell = sheet.cell(
								column=f_col + prev_day_coor[1],
								row=f_row + 2 + 2 * prev_day_coor[0] + 1
							)
							prev_lunar_cell.font = CalendarPrinter.lunar_ft_normal  # unmark 29/12 as a day off

							prev_solar_cell = sheet.cell(
								column=f_col + prev_day_coor[1],
								row=f_row + 2 + 2 * prev_day_coor[0]
							)
							if prev_day_coor[1] == 5:
								prev_solar_cell.font = CalendarPrinter.solar_ft_sat
							elif prev_day_coor[1] == 6:
								prev_solar_cell.font = CalendarPrinter.solar_ft_sun
							else:
								prev_solar_cell.font = CalendarPrinter.solar_ft_weekday

						upper_cell.font = CalendarPrinter.solar_ft_sun
						lower_cell.font = CalendarPrinter.lunar_ft_dayoff

				prev_day_coor = (i, j)
